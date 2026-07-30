from datetime import time
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from puantaj_report import build_report, create_excel_report, read_puantaj_file, time_to_hours


def sample_frame():
    rows = []
    for day in range(1, 8):
        rows.append({
            "sicilno": "00001",
            "Ad": "TEST",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 6, day),
            "NM": time(9) if day <= 4 else time(0),
            "FM": time(3) if day == 6 else time(0),
            "MS": time(9) if day <= 5 else time(0),
            "EM": time(9) if day == 5 else time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    return pd.DataFrame(rows)


def week_with_leave(leave_column: str, leave_code: str | None = None):
    """Hafta içi bir gün izinli; cumartesi mesaisiz."""
    rows = []
    for day in range(1, 8):
        row = {
            "sicilno": "00002",
            "Ad": "IZIN",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 6, day),
            "NM": time(9) if day <= 4 else time(0),
            "FM": time(0),
            "MS": time(9) if day <= 5 else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
            "Kaynak Kod": "",
        }
        if day == 5:
            row["NM"] = time(0)
            row[leave_column] = time(7, 30)
            if leave_code:
                row["Kaynak Kod"] = leave_code
        rows.append(row)
    return pd.DataFrame(rows)


def test_time_to_hours_supports_excel_values():
    assert time_to_hours(time(8, 30)) == 8.5
    assert time_to_hours(0.5) == 12
    assert time_to_hours("09:15:00") == 9.25


def test_report_applies_weekend_transfer_and_sunday_cut():
    result = build_report(sample_frame(), 2026, 6)
    assert result.weekly.loc[0, "FM→NM Aktarım"] == 3
    assert result.weekly.loc[0, "Kalan FM"] == 0
    assert result.weekly.loc[0, "Pazar Durumu"] == "Kesildi"
    assert result.monthly.loc[0, "05 Cu"] == "M"
    assert result.monthly.loc[0, "06 Ct"] == 3.0
    assert result.monthly.loc[0, "07 Pz"] == "Z"


def test_saturday_without_overtime_is_a3():
    frame = sample_frame()
    frame.loc[frame["mesaitarih"] == pd.Timestamp(2026, 6, 6), "FM"] = time(0)
    result = build_report(frame, 2026, 6)
    assert result.monthly.loc[0, "06 Ct"] == "A3"
    assert result.monthly.loc[0, "07 Pz"] == "Z"


def test_protected_leave_does_not_cut_sunday():
    for column, expected_code in (
        ("YIZS", "Y"),
        ("SGKIZS", "R"),
        ("IZS", "Ü"),
    ):
        result = build_report(week_with_leave(column), 2026, 6)
        assert result.weekly.loc[0, "Pazar Durumu"] == "Hak Edildi", column
        assert result.monthly.loc[0, "05 Cu"] == expected_code, column
        assert result.monthly.loc[0, "06 Ct"] == "A3", column
        assert result.monthly.loc[0, "07 Pz"] == "T", column


def test_excel_report_is_created():
    data = create_excel_report(build_report(sample_frame(), 2026, 6))
    assert data.startswith(b"PK")
    assert len(data) > 5000


def test_excel_report_handles_reordered_snapshot_columns():
    result = build_report(sample_frame(), 2026, 6)
    result.monthly = result.monthly[sorted(result.monthly.columns)]

    data = create_excel_report(result)
    workbook = load_workbook(BytesIO(data))
    headers = [cell.value for cell in workbook["Aylık Puantaj"][1]]

    assert headers[:8] == [
        "Sicil No",
        "Personel",
        "Firma",
        "Bölüm",
        "Pozisyon",
        "Görev",
        "Yaka",
        "01 Pt",
    ]


def test_numeric_file_name_does_not_break_reader(tmp_path):
    path = tmp_path / "sample.xlsx"
    sample_frame().to_excel(path, index=False)
    with path.open("rb") as source:
        result = read_puantaj_file(source, filename=123)
    assert len(result) == 7


def test_sakra_horizontal_workbook_is_detected():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026 Puantaj"
    sheet.cell(2, 3, " ADI SOYADI")
    sheet.cell(2, 13, "Yıl")
    sheet.cell(2, 14, "Ay")
    sheet.cell(3, 3, "TEST PERSONEL")
    sheet.cell(3, 13, 2026)
    sheet.cell(3, 14, 7)
    sheet.cell(3, 27, 9)  # 1 Temmuz 2026 Çarşamba, ilk günlük blokta AA.
    sheet.cell(3, 37, 9)  # İlk haftanın normal çalışma sonucu.
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    source = read_puantaj_file(buffer, "sakra.xlsx")
    result = build_report(source, 2026, 7)

    assert len(source) == 1
    assert result.monthly.loc[0, "01 Ça"] == 9


def july_week_with_june_context(*, june_absent: bool = True, june_leave_col: str | None = None):
    """Temmuz 2026 Çarşamba başlar; ISO hafta: 29–30 Haz + 1–5 Tem."""
    rows = []
    # Bağlam: Pazartesi–Salı (Haziran)
    for day, weekday_nm in ((29, 0 if june_absent and not june_leave_col else 9), (30, 0 if june_absent and not june_leave_col else 9)):
        row = {
            "sicilno": "00010",
            "Ad": "BAGLAM",
            "Soyad": "TEST",
            "mesaitarih": pd.Timestamp(2026, 6, day),
            "NM": time(weekday_nm) if weekday_nm else time(0),
            "FM": time(0),
            "MS": time(9),
            "EM": time(9) if june_absent and not june_leave_col else time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
            "Kaynak Kod": "",
        }
        if june_leave_col:
            row["NM"] = time(0)
            row["EM"] = time(0)
            row[june_leave_col] = time(7, 30)
        rows.append(row)

    # Temmuz: Çar–Cum çalış, Cmt mesaisiz, Pz hafta tatili
    for day in range(1, 6):
        is_weekday = day <= 3  # 1 Ça, 2 Pe, 3 Cu
        rows.append({
            "sicilno": "00010",
            "Ad": "BAGLAM",
            "Soyad": "TEST",
            "mesaitarih": pd.Timestamp(2026, 7, day),
            "NM": time(9) if is_weekday else time(0),
            "FM": time(0),
            "MS": time(9) if is_weekday else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
            "Kaynak Kod": "",
        })
    return pd.DataFrame(rows)


def test_context_june_absence_cuts_july_sunday():
    frame = july_week_with_june_context(june_absent=True)
    result = build_report(frame, 2026, 7)

    assert result.monthly.loc[0, "05 Pz"] == "Z*"
    assert result.weekly.loc[0, "Pazar Durumu"] == "Kesildi"
    assert result.weekly.loc[0, "Pazar Kaynağı"] == "Bağlam"
    # Detayda yalnızca Temmuz; Haziran bağlam satırları dönem dışına düşer
    assert result.daily["Tarih"].dt.month.eq(7).all()
    assert len(result.daily) == 5
    # Bağlam NM ve Haziran satırları Temmuz Normal Çalışma'ya eklenmez; yalnızca 3×9
    assert float(result.summary.loc[0, "Normal Çalışma"]) == 27.0
    assert int(result.summary.loc[0, "Devamsızlık (gün)"]) == 0


def test_context_june_protected_leave_keeps_july_sunday():
    frame = july_week_with_june_context(june_absent=False, june_leave_col="YIZS")
    result = build_report(frame, 2026, 7)

    assert result.weekly.loc[0, "Pazar Durumu"] == "Hak Edildi"
    assert result.monthly.loc[0, "05 Pz"] == "T"
    assert result.monthly.loc[0, "04 Ct"] == "A3"
    assert float(result.summary.loc[0, "Normal Çalışma"]) == 27.0
    assert int(result.summary.loc[0, "Yıllık İzin (gün)"]) == 0


def test_download_81_june_context_absence_burns_first_july_sunday():
    """Temmuz 2026: ISO hafta 29–30 Haz + 1–5 Tem; 30 Haz devamsızlık pazarı keser."""
    fixture = Path(__file__).resolve().parent / "fixtures" / "download_81_test.csv"
    source = read_puantaj_file(fixture, fixture.name)
    result = build_report(source, 2026, 7)

    assert result.monthly.loc[0, "05 Pz"] == "Z*"
    first_week = result.weekly[result.weekly["Hafta"].eq("2026-H27")].iloc[0]
    assert first_week["Pazar Durumu"] == "Kesildi"
    assert first_week["Pazar Kaynağı"] == "Bağlam"
    assert result.daily["Tarih"].dt.month.eq(7).all()
