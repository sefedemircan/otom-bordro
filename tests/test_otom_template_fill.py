from datetime import time
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from otom_template_fill import (
    compute_carryover_hours,
    detect_puantaj_sheet,
    fill_otom_template,
    monthly_codes_by_person,
    resolve_day_columns,
    split_holiday_overtime,
    template_week_for_day,
)
from puantaj_report import build_report, prepare_daily, read_puantaj_file, _normalized_text
from test_puantaj_report import sample_frame


WEB_PUBLIC = Path(r"c:\Users\efecan.demircan\Desktop\otom-bordro-web\otom\public")
JULY_TEMPLATE = WEB_PUBLIC / "TEMMUZ OTOM PUANTAJ KOPYA.xlsx"
JULY_CSV = WEB_PUBLIC / "download 2.csv"


def _mini_template(
    year: int = 2026,
    month: int = 6,
    *,
    employees: list[tuple[str, object | None, object | None]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "2026 Puantaj"
    headers = [
        "Ayid",
        "SN",
        " ADI SOYADI",
        "Görevi",
        "ŞUBE",
        "Departman",
        "m/s",
        "Saatlik",
        "Ücret",
        "GİRİŞ",
        "ÇIKIŞ",
        "Yıl",
        "Ay",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(2, col, header)
    ws.cell(2, 23, "Önceki ay son hafta devreden fiili çalışma s.")
    for offset, label in enumerate(("Pt1", "Sa1", "Ça1", "Pe1", "Cu1", "Ct1", "Pz1")):
        ws.cell(1, 24 + offset, offset + 1)
        ws.cell(2, 24 + offset, label)
    ws.cell(2, 31, "1")
    for offset, label in enumerate(("Pt2", "Sa2", "Ça2", "Pe2", "Cu2", "Ct2", "Pz2")):
        ws.cell(1, 43 + offset, 8 + offset)
        ws.cell(2, 43 + offset, label)
    for offset, label in enumerate(("Pt3", "Sa3", "Ça3", "Pe3", "Cu3", "Ct3", "Pz3")):
        ws.cell(2, 62 + offset, label)
    for offset, label in enumerate(("Pt4", "Sa4", "Ça4", "Pe4", "Cu4", "Ct4", "Pz4")):
        ws.cell(2, 81 + offset, label)
    for offset, label in enumerate(("Pt5", "Sa5", "Ça5", "Pe5", "Cu5", "Ct5", "Pz5")):
        ws.cell(2, 100 + offset, label)
    # Holiday manual columns per week (under / over)
    for week, under_col, over_col in (
        (1, 40, 42),
        (2, 59, 61),
        (3, 78, 80),
        (4, 97, 99),
        (5, 116, 118),
    ):
        ws.cell(
            2,
            under_col,
            f"{week}h Hafta Tat+ Resmi Tat F. Ç. 45 i aşmayan kısım (Manuel)",
        )
        ws.cell(
            2,
            over_col,
            f"{week}h Hafta Tat. + R. T. Fazla Ç. 45 i aşan kısım (Manuel)",
        )

    default_employees = employees or [
        ("TEST PERSONEL", None, None),
        ("UNKNOWN PERSON", None, None),
    ]
    for idx, (name, _giris, _cikis) in enumerate(default_employees):
        row = 3 + idx
        ws.cell(row, 3, f"  {name}")
        ws.cell(row, 12, year)
        ws.cell(row, 13, month)

    emp = wb.create_sheet("Çalışan Bilgileri")
    emp.cell(1, 3, " ADI SOYADI")
    emp.cell(1, 11, "GİRİŞ TARİHİ")
    emp.cell(1, 12, "ÇIKIŞ TARİHİ")
    for idx, (name, giris, cikis) in enumerate(default_employees):
        emp.cell(2 + idx, 3, f"  {name}")
        if giris is not None:
            emp.cell(2 + idx, 11, giris)
        if cikis is not None:
            emp.cell(2 + idx, 12, cikis)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_resolve_day_columns_monday_start_june():
    # June 2026 starts Monday → day 1 = Pt1
    wb = load_workbook(BytesIO(_mini_template(2026, 6)))
    ws = detect_puantaj_sheet(wb)
    mapping = resolve_day_columns(ws, 2026, 6)
    assert mapping[1] == 24  # Pt1
    assert mapping[7] == 30  # Pz1
    assert mapping[8] == 43  # Pt2


def test_resolve_day_columns_wednesday_start_july():
    # July 2026 starts Wednesday → day 1 = Ça1, Pt1/Sa1 unused
    wb = load_workbook(BytesIO(_mini_template(2026, 7)))
    ws = detect_puantaj_sheet(wb)
    mapping = resolve_day_columns(ws, 2026, 7)
    assert mapping[1] == 26  # Ça1
    assert mapping[2] == 27  # Pe1
    assert mapping[5] == 30  # July 5 Sunday = Pz1
    assert mapping[6] == 43  # July 6 Monday = Pt2
    assert mapping[31] == 104  # July 31 Friday = Cu5
    assert 24 not in mapping.values()  # Pt1 unused
    assert 25 not in mapping.values()  # Sa1 unused


def test_fill_otom_template_writes_day_codes():
    result = build_report(sample_frame(), 2026, 6)
    filled, stats = fill_otom_template(_mini_template(), result, 2026, 6)
    assert stats.matched == 1
    assert stats.unmatched_template == ["UNKNOWN PERSON"]
    assert stats.filled_cells > 0

    wb = load_workbook(BytesIO(filled))
    ws = detect_puantaj_sheet(wb)
    assert ws.cell(3, 24).value is not None
    assert ws.cell(4, 24).value is None


def test_july_alignment_leaves_pt1_sa1_empty_and_fills_carry():
    rows = []
    # Previous month Mon-Tue that belong to July's first week
    for day, nm in ((29, 9), (30, 9)):
        rows.append({
            "sicilno": "00001",
            "Ad": "TEST",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 6, day),
            "NM": time(nm),
            "FM": time(0),
            "MS": time(9),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    # July days Wed-Sun
    for day in range(1, 8):
        rows.append({
            "sicilno": "00001",
            "Ad": "TEST",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 7, day),
            "NM": time(9) if day <= 5 else time(0),
            "FM": time(0),
            "MS": time(9) if day <= 5 else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    source = pd.DataFrame(rows)
    result = build_report(source, 2026, 7)
    filled, stats = fill_otom_template(
        _mini_template(2026, 7),
        result,
        2026,
        7,
        source_df=source,
    )
    wb = load_workbook(BytesIO(filled))
    ws = detect_puantaj_sheet(wb)

    assert ws.cell(3, 24).value is None  # Pt1 empty
    assert ws.cell(3, 25).value is None  # Sa1 empty
    assert ws.cell(3, 26).value is not None  # Ça1 = July 1
    assert ws.cell(3, 23).value == 18  # 9+9 carry
    assert stats.carry_filled == 1


def test_monthly_codes_normalize_z_star():
    monthly = pd.DataFrame(
        [
            {
                "Personel": "Ali Veli",
                "01 Pt": "Z*",
                "02 Sa": 9.0,
                "03 Ça": "T",
            }
        ]
    )
    by_person = monthly_codes_by_person(monthly)
    assert by_person["ALI VELI"][1] == "Z"
    assert by_person["ALI VELI"][2] == 9
    assert by_person["ALI VELI"][3] == "T"


def test_split_holiday_overtime_example():
    under, over = split_holiday_overtime(43, 4)
    assert under == 2
    assert over == 2
    assert split_holiday_overtime(45, 4) == (0, 4)
    assert split_holiday_overtime(40, 3) == (3, 0)


def test_july_15_is_week_3_and_forced_b_with_manual_split():
    rows = []
    # Week 3 days for July 2026: 13-19 (Wed start month → week3)
    assert template_week_for_day(2026, 7, 15) == 3
    for day, nm, fm, rm, rm_desc in (
        (13, 9, 0, 0, ""),
        (14, 9, 0, 0, ""),
        (15, 4, 0, 0, "15.Tem"),  # holiday work 4h; week so far 18 → room 27
        (16, 9, 0, 0, ""),
        (17, 9, 0, 0, ""),
        (18, 0, 0, 0, ""),
        (19, 0, 0, 0, ""),
    ):
        rows.append({
            "sicilno": "00001",
            "Ad": "TEST",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 7, day),
            "NM": time(nm) if nm else time(0),
            "FM": time(fm) if fm else time(0),
            "MS": time(9) if day <= 17 else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(rm) if rm else time(0),
            "RM Açıklama": rm_desc,
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    # Second person: week hours 43 equivalent via 9*4 + 7 = 43 before holiday? 
    # Use explicit: 13-14:9+9=18, 16-17:9+9=18, total 36 without holiday; holiday 4 → all under
    source = pd.DataFrame(rows)
    # Add high-hour person for 43+4 split
    extra = []
    for day, nm in ((13, 9), (14, 9), (15, 4), (16, 9), (17, 9), (18, 7), (19, 0)):
        extra.append({
            "sicilno": "00002",
            "Ad": "HIGH",
            "Soyad": "HOURS",
            "mesaitarih": pd.Timestamp(2026, 7, day),
            "NM": time(nm) if nm else time(0),
            "FM": time(0),
            "MS": time(9) if nm else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "RM Açıklama": "15.Tem" if day == 15 else "",
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    source = pd.concat([source, pd.DataFrame(extra)], ignore_index=True)
    # Expand mini template names
    tpl = _mini_template(2026, 7)
    wb = load_workbook(BytesIO(tpl))
    ws = detect_puantaj_sheet(wb)
    ws.cell(5, 3, "  HIGH HOURS")
    ws.cell(5, 12, 2026)
    ws.cell(5, 13, 7)
    buf = BytesIO()
    wb.save(buf)
    tpl = buf.getvalue()

    result = build_report(source, 2026, 7)
    assert result.monthly.loc[result.monthly["Personel"].eq("TEST PERSONEL"), "15 Ça"].iloc[0] == "B"

    filled, stats = fill_otom_template(tpl, result, 2026, 7, source_df=source)
    out = load_workbook(BytesIO(filled))
    sheet = detect_puantaj_sheet(out)
    # TEST PERSONEL row 3: July 15 = Ça3 col 64
    assert sheet.cell(3, 64).value == "B"
    # week hours excluding holiday: 9*4=36, holiday 4 → under 4, over 0
    assert sheet.cell(3, 78).value == 4
    assert sheet.cell(3, 80).value in (None, 0, "")
    # HIGH HOURS: week excl holiday = 9+9+9+9+7=43, holiday 4 → 2 / 2
    assert sheet.cell(5, 64).value == "B"
    assert sheet.cell(5, 78).value == 2
    assert sheet.cell(5, 80).value == 2
    assert stats.holiday_manual_filled >= 1


def test_employment_dates_fill_x_before_hire_and_after_exit():
    rows = []
    for day in range(1, 32):
        rows.append({
            "sicilno": "00001",
            "Ad": "TEST",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 7, day),
            "NM": time(9) if day < 6 else time(0),
            "FM": time(0),
            "MS": time(9) if day < 6 else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    source = pd.DataFrame(rows)
    result = build_report(source, 2026, 7)
    tpl = _mini_template(
        2026,
        7,
        employees=[
            ("TEST PERSONEL", pd.Timestamp(2026, 7, 16), pd.Timestamp(2026, 7, 19)),
            ("UNKNOWN PERSON", None, None),
        ],
    )
    filled, _stats = fill_otom_template(tpl, result, 2026, 7, source_df=source)
    wb = load_workbook(BytesIO(filled))
    ws = detect_puantaj_sheet(wb)
    mapping = resolve_day_columns(ws, 2026, 7)
    # Before hire (1..15) → X
    assert ws.cell(3, mapping[15]).value == "X"
    # Hire day and exit day inclusive stay non-X from employment rule
    assert ws.cell(3, mapping[16]).value != "X"
    assert ws.cell(3, mapping[19]).value != "X"
    # After exit (20+) → X
    assert ws.cell(3, mapping[20]).value == "X"
    assert ws.cell(3, mapping[31]).value == "X"


def test_holiday_hours_from_punch_when_nm_zero():
    """Meyer resmi tatilde NM=0 bırakıp giriş/çıkış basmış olabilir."""
    rows = []
    for day, nm, giris, cikis in (
        (13, 9, "08:00:00", "18:15:00"),
        (14, 9, "08:00:00", "18:15:00"),
        (15, 0, "07:40:00", "18:15:00"),  # holiday work via punch only
        (16, 9, "08:00:00", "18:15:00"),
        (17, 9, "08:00:00", "18:15:00"),
        (18, 0, None, None),
        (19, 0, None, None),
    ):
        rows.append({
            "sicilno": "00003",
            "Ad": "PUNCH",
            "Soyad": "WORKER",
            "mesaitarih": pd.Timestamp(2026, 7, day),
            "NM": time(nm) if nm else time(0),
            "FM": time(0),
            "MS": time(9) if day <= 17 else time(0),
            "EM": time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(10) if day == 15 else time(0),
            "RM Açıklama": "15.Tem" if day == 15 else "",
            "Giriş": giris,
            "Çıkış": cikis,
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    source = pd.DataFrame(rows)
    result = build_report(source, 2026, 7)
    tpl = _mini_template(
        2026,
        7,
        employees=[("PUNCH WORKER", pd.Timestamp(2026, 1, 1), None)],
    )
    filled, stats = fill_otom_template(tpl, result, 2026, 7, source_df=source)
    wb = load_workbook(BytesIO(filled))
    ws = detect_puantaj_sheet(wb)
    # week hours excl holiday ~= 36, punch net ~= 9.33 → room 9 → under 9 / over 0.33
    assert ws.cell(3, 64).value == "B"
    assert float(ws.cell(3, 78).value) == 9
    assert abs(float(ws.cell(3, 80).value) - 0.33) < 0.01
    assert stats.holiday_manual_filled >= 1


def test_july_real_files_smoke():
    if not JULY_TEMPLATE.exists() or not JULY_CSV.exists():
        return

    df = read_puantaj_file(JULY_CSV.open("rb"), filename=JULY_CSV.name)
    result = build_report(df, 2026, 7)
    filled, stats = fill_otom_template(
        JULY_TEMPLATE.read_bytes(),
        result,
        2026,
        7,
        source_df=df,
    )
    assert stats.sheet_name.strip().endswith("Puantaj")
    assert stats.matched >= 200
    assert stats.filled_cells >= 5000
    assert stats.carry_filled >= 1
    assert filled.startswith(b"PK")

    wb = load_workbook(BytesIO(filled))
    ws = detect_puantaj_sheet(wb)
    src = {_normalized_text(p) for p in result.monthly["Personel"]}
    matched_row = None
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row_idx, 3).value
        if name and _normalized_text(name) in src:
            matched_row = row_idx
            break
    assert matched_row is not None
    assert ws.cell(matched_row, 24).value in (None, "")
    assert ws.cell(matched_row, 25).value in (None, "")
    assert ws.cell(matched_row, 26).value is not None  # Ça1
