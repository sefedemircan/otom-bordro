"""OTOM aylık puantaj şablonunu rapor çıktısıyla doldurur."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from puantaj_report import (
    WEEKLY_MAX_HOURS,
    ReportResult,
    _column_key,
    _is_public_holiday_date,
    _normalized_text,
    prepare_daily,
)


DAY_SLOT_PATTERN = re.compile(
    r"^(PT|SA|CA|PE|CU|CT|PZ)(\d+)$",
    re.IGNORECASE,
)
DAY_COLUMN_PATTERN = re.compile(r"^(0[1-9]|[12]\d|3[01]) (Pt|Sa|Ça|Pe|Cu|Ct|Pz)$")


@dataclass
class FillStats:
    matched: int = 0
    unmatched_template: list[str] = field(default_factory=list)
    unmatched_source: list[str] = field(default_factory=list)
    filled_cells: int = 0
    carry_filled: int = 0
    holiday_manual_filled: int = 0
    sheet_name: str = ""
    year: int | None = None
    month: int | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "matched": self.matched,
            "unmatched_template": self.unmatched_template,
            "unmatched_source": self.unmatched_source,
            "filled_cells": self.filled_cells,
            "carry_filled": self.carry_filled,
            "holiday_manual_filled": self.holiday_manual_filled,
            "sheet_name": self.sheet_name,
            "year": self.year,
            "month": self.month,
        }


def _header_row(ws: Worksheet) -> int:
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            if _column_key(ws.cell(row, col).value) == "ADISOYADI":
                return row
    raise ValueError("Şablonda 'ADI SOYADI' başlığı bulunamadı.")


def _find_name_column(ws: Worksheet, header_row: int) -> int:
    for col in range(1, ws.max_column + 1):
        if _column_key(ws.cell(header_row, col).value) == "ADISOYADI":
            return col
    raise ValueError("Şablonda 'ADI SOYADI' kolonu bulunamadı.")


def _find_carry_column(ws: Worksheet, header_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        key = _column_key(ws.cell(header_row, col).value)
        if "ONCEKI" in key and "DEVREDEN" in key:
            return col
    return None


def _sheet_day_slot_count(ws: Worksheet) -> int:
    try:
        header_row = _header_row(ws)
    except ValueError:
        return 0
    return sum(
        1
        for col in range(1, ws.max_column + 1)
        if _slot_key(ws.cell(header_row, col).value)
    )


def detect_puantaj_sheet(wb) -> Worksheet:
    candidates: list[tuple[int, Worksheet]] = []
    for ws in wb.worksheets:
        slot_count = _sheet_day_slot_count(ws)
        if slot_count == 0:
            continue
        bonus = 100 if "PUANTAJ" in _normalized_text(ws.title) else 0
        candidates.append((slot_count + bonus, ws))
    if not candidates:
        raise ValueError("OTOM puantaj şablonu sheet'i bulunamadı (ADI SOYADI + Pt1… gerekli).")
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _slot_key(value: object) -> tuple[int, int] | None:
    text = _normalized_text(value).replace(" ", "")
    match = DAY_SLOT_PATTERN.fullmatch(text)
    if not match:
        return None
    weekday_token, week_text = match.group(1).upper(), match.group(2)
    weekday_order = {"PT": 0, "SA": 1, "CA": 2, "PE": 3, "CU": 4, "CT": 5, "PZ": 6}
    weekday = weekday_order.get(weekday_token)
    if weekday is None:
        return None
    return int(week_text), weekday


def resolve_week_starts(ws: Worksheet) -> dict[int, int]:
    """1-based week index → first day-slot column (PtN)."""
    header_row = _header_row(ws)
    week_starts: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        slot = _slot_key(ws.cell(header_row, col).value)
        if not slot:
            continue
        week_index, weekday = slot
        if weekday == 0:
            week_starts[week_index] = col
        elif week_index not in week_starts:
            week_starts[week_index] = col - weekday
    if not week_starts:
        raise ValueError("Şablonda haftalık gün kolonları (Pt1, Sa1, …) bulunamadı.")
    return week_starts


def resolve_day_columns(ws: Worksheet, year: int, month: int) -> dict[int, int]:
    """Ay günü (1..N) → Excel kolon index.

    Ayın ilk gününün hafta içi konumuna göre hizalanır:
    Temmuz 2026 Çarşamba başlar → gün 1 = Ça1, Pt1/Sa1 boş kalır.
    """
    week_starts = resolve_week_starts(ws)
    first_weekday = int(pd.Timestamp(year=year, month=month, day=1).dayofweek)
    days_in_month = int(pd.Timestamp(year=year, month=month, day=1).days_in_month)
    mapping: dict[int, int] = {}
    for day in range(1, days_in_month + 1):
        slot = first_weekday + day - 1
        week_index_0, weekday = divmod(slot, 7)
        week_index = week_index_0 + 1
        start = week_starts.get(week_index)
        if start is None:
            continue
        mapping[day] = start + weekday
    if not mapping:
        raise ValueError("Şablon gün kolonları dönem günleriyle eşleştirilemedi.")
    return mapping


def _hours_value(row: pd.Series) -> float:
    for left, right in (
        ("NM Güncel_h", "FM Güncel_h"),
        ("NM_h", "FM_h"),
        ("Normal Çalışma", "Fazla Mesai"),
    ):
        if left in row.index or right in row.index:
            left_val = row.get(left, 0)
            right_val = row.get(right, 0)
            try:
                return float(left_val or 0) + float(right_val or 0)
            except (TypeError, ValueError):
                pass
    kod = row.get("Kod")
    if isinstance(kod, (int, float)) and not pd.isna(kod):
        return float(kod) if float(kod) > 0 else 0.0
    if isinstance(kod, str):
        text = kod.strip().replace(",", ".")
        try:
            number = float(text)
        except ValueError:
            return 0.0
        return number if number > 0 else 0.0
    return 0.0


def compute_carryover_hours(daily_all: pd.DataFrame, year: int, month: int) -> dict[str, float]:
    """Ayın ilk haftasındaki önceki-ay günlerinin fiili çalışma toplamı."""
    if daily_all is None or daily_all.empty or "Tarih" not in daily_all.columns:
        return {}
    period_start = pd.Timestamp(year=year, month=month, day=1)
    first_weekday = int(period_start.dayofweek)
    if first_weekday == 0:
        return {}

    carry_start = period_start - pd.Timedelta(days=first_weekday)
    carry_end = period_start - pd.Timedelta(days=1)
    window = daily_all[
        (daily_all["Tarih"] >= carry_start) & (daily_all["Tarih"] <= carry_end)
    ].copy()
    if window.empty:
        return {}

    totals: dict[str, float] = {}
    for _, row in window.iterrows():
        name = _normalized_text(row.get("Personel", ""))
        if not name:
            continue
        hours = _hours_value(row)
        if hours <= 0:
            continue
        totals[name] = round(totals.get(name, 0.0) + hours, 2)
    return totals


def carryover_hours_from_source(source_df: pd.DataFrame, year: int, month: int) -> dict[str, float]:
    daily_all, _, _ = prepare_daily(source_df)
    return compute_carryover_hours(daily_all, year, month)


def carryover_hours_from_upload_snapshots(upload_id: str, year: int, month: int) -> dict[str, float]:
    """Report upload'ları ham satır tutmaz; önceki ay snapshot daily'sinden carry hesapla."""
    from api.services.report_snapshots import load_report_result_from_snapshot
    from api.services.supabase import SupabaseError

    period_start = pd.Timestamp(year=year, month=month, day=1)
    first_weekday = int(period_start.dayofweek)
    if first_weekday == 0:
        return {}

    carry_start = period_start - pd.Timedelta(days=first_weekday)
    carry_end = period_start - pd.Timedelta(days=1)
    months_needed = sorted({
        (int(carry_start.year), int(carry_start.month)),
        (int(carry_end.year), int(carry_end.month)),
    })

    frames: list[pd.DataFrame] = []
    for prev_year, prev_month in months_needed:
        if (prev_year, prev_month) == (year, month):
            continue
        try:
            prev_result = load_report_result_from_snapshot(upload_id, prev_year, prev_month)
        except SupabaseError:
            continue
        if prev_result.daily is None or prev_result.daily.empty:
            continue
        frames.append(prev_result.daily.copy())

    if not frames:
        return {}

    daily = pd.concat(frames, ignore_index=True)
    if "Tarih" in daily.columns:
        daily["Tarih"] = pd.to_datetime(daily["Tarih"], format="mixed", dayfirst=True, errors="coerce")
    return compute_carryover_hours(daily, year, month)


def _cell_value_for_template(raw: object) -> object | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        number = float(raw)
        if number <= 0:
            return None
        return int(number) if number.is_integer() else round(number, 2)
    text = str(raw).strip()
    if not text or text.lower() in {"nan", "none", "nat"}:
        return None
    if text == "Z*":
        return "Z"
    numeric = text.replace(",", ".")
    try:
        number = float(numeric)
    except ValueError:
        return text
    if number <= 0:
        return None
    return int(number) if number.is_integer() else round(number, 2)


def monthly_codes_by_person(monthly: pd.DataFrame) -> dict[str, dict[int, object]]:
    if monthly is None or monthly.empty:
        return {}
    day_columns = [column for column in monthly.columns if DAY_COLUMN_PATTERN.fullmatch(str(column))]
    by_person: dict[str, dict[int, object]] = {}
    for row in monthly.to_dict(orient="records"):
        name = _normalized_text(row.get("Personel", ""))
        if not name:
            continue
        day_map: dict[int, object] = {}
        for column in day_columns:
            day = int(str(column)[:2])
            value = _cell_value_for_template(row.get(column))
            if value is not None:
                day_map[day] = value
        by_person[name] = day_map
    return by_person


def _all_day_slot_columns(ws: Worksheet) -> list[int]:
    header_row = _header_row(ws)
    return [
        col
        for col in range(1, ws.max_column + 1)
        if _slot_key(ws.cell(header_row, col).value)
    ]


def resolve_holiday_manual_columns(ws: Worksheet) -> dict[int, tuple[int | None, int | None]]:
    """Hafta no → (45'i aşmayan manuel kolon, 45'i aşan manuel kolon)."""
    header_row = _header_row(ws)
    under: dict[int, int] = {}
    over: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _column_key(ws.cell(header_row, col).value)
        match = re.match(r"^(\d+)H", key)
        if not match or "MANUEL" not in key:
            continue
        week = int(match.group(1))
        if "ASMAYAN" in key and ("HAFTATAT" in key or "RESMITAT" in key):
            under[week] = col
        elif "ASAN" in key and "ASMAYAN" not in key and ("HAFTATAT" in key or "RT" in key):
            over[week] = col
    weeks = sorted(set(under) | set(over))
    return {week: (under.get(week), over.get(week)) for week in weeks}


def template_week_for_day(year: int, month: int, day: int) -> int:
    first_weekday = int(pd.Timestamp(year=year, month=month, day=1).dayofweek)
    return (first_weekday + day - 1) // 7 + 1


def days_in_template_week(year: int, month: int, week_index: int) -> list[int]:
    days_in_month = int(pd.Timestamp(year=year, month=month, day=1).days_in_month)
    return [
        day
        for day in range(1, days_in_month + 1)
        if template_week_for_day(year, month, day) == week_index
    ]


def detect_holiday_days(daily: pd.DataFrame, year: int, month: int) -> set[int]:
    days_in_month = int(pd.Timestamp(year=year, month=month, day=1).days_in_month)
    holidays = {
        day
        for day in range(1, days_in_month + 1)
        if _is_public_holiday_date(pd.Timestamp(year=year, month=month, day=day))
    }
    if daily is None or daily.empty or "Tarih" not in daily.columns:
        return holidays
    frame = daily.copy()
    frame["Tarih"] = pd.to_datetime(frame["Tarih"], format="mixed", dayfirst=True, errors="coerce")
    frame = frame.dropna(subset=["Tarih"])
    frame = frame[
        (frame["Tarih"].dt.year == year) & (frame["Tarih"].dt.month == month)
    ]
    if "Durum" in frame.columns:
        holidays.update(
            int(day)
            for day in frame.loc[frame["Durum"].eq("RESMI_TATIL"), "Tarih"].dt.day.tolist()
        )
    return holidays


def apply_holiday_codes(
    source_by_name: dict[str, dict[int, object]],
    holiday_days: set[int],
) -> None:
    for day_map in source_by_name.values():
        for day in holiday_days:
            if day_map.get(day) == "X":
                continue
            day_map[day] = "B"


def _parse_excel_date(value: object) -> pd.Timestamp | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        ts = pd.Timestamp(value)
    elif isinstance(value, date):
        ts = pd.Timestamp(value)
    else:
        ts = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if ts is None or pd.isna(ts):
        return None
    # Excel empty date cells sometimes become 00:00:00 / 1899-12-30.
    if ts.year < 1905:
        return None
    return ts.normalize()


def _find_employee_sheet(wb):
    for ws in wb.worksheets:
        title_key = _column_key(ws.title)
        if "CALISAN" in title_key and "BILGI" in title_key:
            return ws
        # Header-based detection
        for row in range(1, min(4, ws.max_row + 1)):
            keys = {_column_key(ws.cell(row, col).value) for col in range(1, min(25, ws.max_column + 1))}
            if "ADISOYADI" in keys and "GIRISTARIHI" in keys:
                return ws
    return None


def load_employment_windows(wb) -> dict[str, tuple[pd.Timestamp | None, pd.Timestamp | None]]:
    """ADI SOYADI → (giriş, çıkış) — Çalışan Bilgileri sheet'inden."""
    ws = _find_employee_sheet(wb)
    if ws is None:
        return {}

    header_row = None
    for row in range(1, min(5, ws.max_row + 1)):
        for col in range(1, min(25, ws.max_column + 1)):
            if _column_key(ws.cell(row, col).value) == "ADISOYADI":
                header_row = row
                break
        if header_row is not None:
            break
    if header_row is None:
        return {}

    name_col = giris_col = cikis_col = None
    for col in range(1, ws.max_column + 1):
        key = _column_key(ws.cell(header_row, col).value)
        if key == "ADISOYADI":
            name_col = col
        elif key == "GIRISTARIHI":
            giris_col = col
        elif key == "CIKISTARIHI":
            cikis_col = col
    if name_col is None:
        return {}

    windows: dict[str, tuple[pd.Timestamp | None, pd.Timestamp | None]] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row_idx, name_col).value
        if raw_name is None or not str(raw_name).strip():
            continue
        key = _normalized_text(raw_name)
        giris = _parse_excel_date(ws.cell(row_idx, giris_col).value) if giris_col else None
        cikis = _parse_excel_date(ws.cell(row_idx, cikis_col).value) if cikis_col else None
        windows[key] = (giris, cikis)
    return windows


def apply_employment_outside_codes(
    source_by_name: dict[str, dict[int, object]],
    year: int,
    month: int,
    employment: dict[str, tuple[pd.Timestamp | None, pd.Timestamp | None]],
) -> None:
    """Giriş öncesi / çıkış sonrası günleri X (sayılmayan) yap."""
    days_in_month = int(pd.Timestamp(year=year, month=month, day=1).days_in_month)
    for name, day_map in source_by_name.items():
        giris, cikis = employment.get(name, (None, None))
        if giris is None and cikis is None:
            continue
        for day in range(1, days_in_month + 1):
            current = pd.Timestamp(year=year, month=month, day=day)
            if giris is not None and current < giris:
                day_map[day] = "X"
            elif cikis is not None and current > cikis:
                day_map[day] = "X"


def _raw_work_hours(row: pd.Series) -> float:
    for left, right in (("NM_h", "FM_h"), ("NM Güncel_h", "FM Güncel_h")):
        if left in row.index or right in row.index:
            try:
                return float(row.get(left, 0) or 0) + float(row.get(right, 0) or 0)
            except (TypeError, ValueError):
                pass
    return _hours_value(row)


def split_holiday_overtime(weekly_hours: float, holiday_hours: float) -> tuple[float, float]:
    if holiday_hours <= 0:
        return 0.0, 0.0
    room = max(0.0, WEEKLY_MAX_HOURS - max(0.0, weekly_hours))
    under = min(holiday_hours, room)
    over = max(0.0, holiday_hours - under)
    return round(under, 2), round(over, 2)


def holiday_manual_hours_by_person(
    daily: pd.DataFrame,
    year: int,
    month: int,
    holiday_days: set[int],
    carry_by_name: dict[str, float] | None = None,
    employment: dict[str, tuple[pd.Timestamp | None, pd.Timestamp | None]] | None = None,
) -> dict[str, dict[int, tuple[float, float]]]:
    """person → week → (aşmayan, aşan) saatleri."""
    if not holiday_days or daily is None or daily.empty:
        return {}

    frame = daily.copy()
    frame["Tarih"] = pd.to_datetime(frame["Tarih"], format="mixed", dayfirst=True, errors="coerce")
    frame = frame.dropna(subset=["Tarih"])
    frame = frame[
        (frame["Tarih"].dt.year == year) & (frame["Tarih"].dt.month == month)
    ].copy()
    if frame.empty:
        return {}

    frame["day"] = frame["Tarih"].dt.day.astype(int)
    frame["week"] = frame["day"].map(lambda day: template_week_for_day(year, month, int(day)))
    frame["work_h"] = frame.apply(_raw_work_hours, axis=1)
    frame["is_holiday"] = frame["day"].isin(holiday_days)
    holiday_weeks = {
        template_week_for_day(year, month, day) for day in holiday_days
    }
    employment = employment or {}

    def _in_employment(person: str, day: int) -> bool:
        giris, cikis = employment.get(person, (None, None))
        current = pd.Timestamp(year=year, month=month, day=day)
        if giris is not None and current < giris:
            return False
        if cikis is not None and current > cikis:
            return False
        return True

    result: dict[str, dict[int, tuple[float, float]]] = {}
    for person, group in frame.groupby(frame["Personel"].map(_normalized_text)):
        if not person:
            continue
        week_map: dict[int, tuple[float, float]] = {}
        for week in holiday_weeks:
            week_rows = group[group["week"] == week].copy()
            if week_rows.empty:
                continue
            week_rows = week_rows[week_rows["day"].map(lambda day: _in_employment(person, int(day)))]
            if week_rows.empty:
                continue
            weekly_hours = float(week_rows.loc[~week_rows["is_holiday"], "work_h"].sum())
            if week == 1 and carry_by_name:
                weekly_hours += float(carry_by_name.get(person, 0) or 0)
            holiday_hours = float(week_rows.loc[week_rows["is_holiday"], "work_h"].sum())
            under, over = split_holiday_overtime(weekly_hours, holiday_hours)
            if under > 0 or over > 0:
                week_map[week] = (under, over)
        if week_map:
            result[person] = week_map
    return result


def fill_otom_template(
    template_bytes: bytes,
    result: ReportResult,
    year: int,
    month: int,
    source_df: pd.DataFrame | None = None,
    carry_by_name: dict[str, float] | None = None,
) -> tuple[bytes, FillStats]:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = detect_puantaj_sheet(wb)
    header_row = _header_row(ws)
    name_col = _find_name_column(ws, header_row)
    carry_col = _find_carry_column(ws, header_row)
    day_slot_columns = _all_day_slot_columns(ws)
    day_columns = resolve_day_columns(ws, year, month)
    holiday_manual_cols = resolve_holiday_manual_columns(ws)
    source_by_name = monthly_codes_by_person(result.monthly)
    employment = load_employment_windows(wb)
    if carry_by_name is not None:
        resolved_carry = carry_by_name
    elif source_df is not None:
        resolved_carry = carryover_hours_from_source(source_df, year, month)
    else:
        resolved_carry = {}

    holiday_days = detect_holiday_days(result.daily, year, month)
    apply_holiday_codes(source_by_name, holiday_days)
    # Giriş/çıkış X kodu tatil B'sinin üzerine yazılır.
    apply_employment_outside_codes(source_by_name, year, month, employment)
    holiday_manual = holiday_manual_hours_by_person(
        result.daily,
        year,
        month,
        holiday_days,
        carry_by_name=resolved_carry,
        employment=employment,
    )
    used_source: set[str] = set()

    stats = FillStats(sheet_name=ws.title, year=year, month=month)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row_idx, name_col).value
        if raw_name is None or not str(raw_name).strip():
            continue
        key = _normalized_text(raw_name)
        day_map = source_by_name.get(key)
        if day_map is None:
            stats.unmatched_template.append(str(raw_name).strip())
            continue
        used_source.add(key)
        stats.matched += 1
        # Clear stale manual values (e.g. Pt1/Sa1 when month starts mid-week).
        # openpyxl ignores value=None in cell(); assign via .value.
        for col_idx in day_slot_columns:
            ws.cell(row_idx, col_idx).value = None
        for day, col_idx in day_columns.items():
            if day not in day_map:
                continue
            ws.cell(row_idx, col_idx).value = day_map[day]
            stats.filled_cells += 1
        if carry_col is not None:
            giris, _cikis = employment.get(key, (None, None))
            carry_hours = resolved_carry.get(key)
            # Ay ortası girişte önceki aydan carry yazılmaz.
            if giris is not None and giris > pd.Timestamp(year=year, month=month, day=1):
                carry_hours = 0
            if carry_hours and carry_hours > 0:
                value = int(carry_hours) if float(carry_hours).is_integer() else carry_hours
                ws.cell(row_idx, carry_col).value = value
                stats.carry_filled += 1
            else:
                ws.cell(row_idx, carry_col).value = 0

        person_holiday = holiday_manual.get(key, {})
        for week, (under_col, over_col) in holiday_manual_cols.items():
            under, over = person_holiday.get(week, (0.0, 0.0))
            if under_col is not None:
                ws.cell(row_idx, under_col).value = (
                    int(under) if float(under).is_integer() else under
                ) if under > 0 else 0
            if over_col is not None:
                ws.cell(row_idx, over_col).value = (
                    int(over) if float(over).is_integer() else over
                ) if over > 0 else 0
            if under > 0 or over > 0:
                stats.holiday_manual_filled += 1

    stats.unmatched_source = sorted(name for name in source_by_name if name not in used_source)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), stats
